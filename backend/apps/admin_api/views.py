import logging
from urllib.parse import urlparse

from django.contrib.auth import authenticate
from django.core.cache import cache
from django.db.models import Sum, Count, Q, Max
from django.db.models.functions import TruncMonth, TruncDate
from django.utils import timezone
from datetime import timedelta
from rest_framework import generics, status
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import serializers
from rest_framework.throttling import ScopedRateThrottle
from rest_framework_simplejwt.tokens import RefreshToken

from .permissions import IsAdminUser, IsAdminOrVendor
from .models import AuditLog
from .audit import log_action

logger = logging.getLogger(__name__)

# ── Security: Account lockout constants ──────────────────────────────────────
LOCKOUT_THRESHOLD = 5  # failed attempts before lockout
LOCKOUT_DURATION = 900  # 15 minutes in seconds

# ── Security: Valid ISO 3166-1 alpha-2 country codes ─────────────────────────
VALID_COUNTRY_CODES = frozenset({
    "AD", "AE", "AF", "AG", "AI", "AL", "AM", "AO", "AQ", "AR", "AS", "AT",
    "AU", "AW", "AX", "AZ", "BA", "BB", "BD", "BE", "BF", "BG", "BH", "BI",
    "BJ", "BL", "BM", "BN", "BO", "BR", "BS", "BT", "BV", "BW", "BY", "BZ",
    "CA", "CC", "CD", "CF", "CG", "CH", "CI", "CK", "CL", "CM", "CN", "CO",
    "CR", "CU", "CV", "CW", "CX", "CY", "CZ", "DE", "DJ", "DK", "DM", "DO",
    "DZ", "EC", "EE", "EG", "EH", "ER", "ES", "ET", "FI", "FJ", "FK", "FM",
    "FO", "FR", "GA", "GB", "GD", "GE", "GF", "GG", "GH", "GI", "GL", "GM",
    "GN", "GP", "GQ", "GR", "GS", "GT", "GU", "GW", "GY", "HK", "HM", "HN",
    "HR", "HT", "HU", "ID", "IE", "IL", "IM", "IN", "IO", "IQ", "IR", "IS",
    "IT", "JE", "JM", "JO", "JP", "KE", "KG", "KH", "KI", "KM", "KN", "KP",
    "KR", "KW", "KY", "KZ", "LA", "LB", "LC", "LI", "LK", "LR", "LS", "LT",
    "LU", "LV", "LY", "MA", "MC", "MD", "ME", "MF", "MG", "MH", "MK", "ML",
    "MM", "MN", "MO", "MP", "MQ", "MR", "MS", "MT", "MU", "MV", "MW", "MX",
    "MY", "MZ", "NA", "NC", "NE", "NF", "NG", "NI", "NL", "NO", "NP", "NR",
    "NU", "NZ", "OM", "PA", "PE", "PF", "PG", "PH", "PK", "PL", "PM", "PN",
    "PR", "PS", "PT", "PW", "PY", "QA", "RE", "RO", "RS", "RU", "RW", "SA",
    "SB", "SC", "SD", "SE", "SG", "SH", "SI", "SJ", "SK", "SL", "SM", "SN",
    "SO", "SR", "SS", "ST", "SV", "SX", "SY", "SZ", "TC", "TD", "TF", "TG",
    "TH", "TJ", "TK", "TL", "TM", "TN", "TO", "TR", "TT", "TV", "TW", "TZ",
    "UA", "UG", "UM", "US", "UY", "UZ", "VA", "VC", "VE", "VG", "VI", "VN",
    "VU", "WF", "WS", "YE", "YT", "ZA", "ZM", "ZW",
})


def _login_lockout_key(email: str) -> str:
    return f"login_lockout:{email.strip().lower()}"


def _login_attempts_key(email: str) -> str:
    return f"login_attempts:{email.strip().lower()}"


def _check_lockout(email: str) -> bool:
    """Returns True if account is currently locked out."""
    return bool(cache.get(_login_lockout_key(email)))


def _record_failed_login(email: str) -> int:
    """Record a failed login attempt. Returns current count. Triggers lockout if threshold hit."""
    key = _login_attempts_key(email)
    attempts = int(cache.get(key, 0) or 0) + 1
    cache.set(key, attempts, timeout=LOCKOUT_DURATION)
    if attempts >= LOCKOUT_THRESHOLD:
        cache.set(_login_lockout_key(email), True, timeout=LOCKOUT_DURATION)
    return attempts


def _clear_failed_logins(email: str) -> None:
    """Clear failed login counter on successful auth."""
    cache.delete(_login_attempts_key(email))
    cache.delete(_login_lockout_key(email))


class AdminNoPaginationMixin:
    pagination_class = None


class AdminAuthThrottle(ScopedRateThrottle):
    scope = "admin_auth"


# ── Admin login ────────────────────────────────────────────────────────────────

class AdminLoginView(APIView):
    """Login endpoint exclusively for staff and vendor users."""
    permission_classes = [AllowAny]
    throttle_classes = [AdminAuthThrottle]

    def post(self, request):
        email = request.data.get("email", "").strip().lower()
        password = request.data.get("password", "")

        # Security fix #9: Account lockout after repeated failures
        if _check_lockout(email):
            return Response(
                {"detail": "Account temporarily locked due to too many failed attempts. Try again later."},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
            )

        user = authenticate(request, username=email, password=password)
        if not user or not user.is_active:
            _record_failed_login(email)
            return Response({"detail": "Invalid credentials."}, status=status.HTTP_401_UNAUTHORIZED)

        is_vendor = hasattr(user, "vendor_profile")
        if not user.is_staff and not is_vendor:
            return Response({"detail": "Access denied. Vendor or admin account required."}, status=status.HTTP_403_FORBIDDEN)

        _clear_failed_logins(email)
        refresh = RefreshToken.for_user(user)

        vendor_data = None
        if is_vendor:
            v = user.vendor_profile
            vendor_data = {"id": v.id, "name": v.name, "slug": v.slug, "logo_url": v.logo_url}

        return Response({
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "user": {
                "id": str(user.id),
                "email": user.email,
                "name": user.name,
                "role": user.role,
                "is_staff": user.is_staff,
                "vendor": vendor_data,
            },
        })


# ── Serializers ──────────────────────────────────────────────────────────────

class AuditLogSerializer(serializers.ModelSerializer):
    admin_email = serializers.CharField(source="admin_user.email", read_only=True, allow_null=True)
    summary = serializers.SerializerMethodField()
    category = serializers.SerializerMethodField()

    class Meta:
        model = AuditLog
        fields = ("id", "admin_email", "action", "target_type", "target_id", "detail", "summary", "category", "timestamp")

    def get_summary(self, obj):
        d = obj.detail or {}
        if obj.action == "order_created":
            return f"Order {d.get('order_number', obj.target_id)} — {d.get('total', '')} {d.get('currency', '')}".strip()
        if obj.action == "order_status_change":
            return f"Order status → {d.get('new_status', '')}"
        if obj.action == "payment_received":
            return f"Payment {d.get('payment_ref', '')} — {d.get('amount', '')} {d.get('currency', '')}".strip()
        if obj.action == "settings_update":
            return f"Updated settings: {', '.join(d.get('keys', []))}"
        if obj.action == "vendor_ops_update":
            return f"Vendor {d.get('vendor_slug', '')} settings updated"
        if obj.action in ("create", "update", "delete") and obj.target_type == "page_section":
            return f"{obj.action.title()} section {d.get('section_key', obj.target_id)}"
        return f"{obj.action} on {obj.target_type} #{obj.target_id}"

    def get_category(self, obj):
        t = obj.target_type.lower()
        if t == "order" or obj.action.startswith("order"):
            return "Order"
        if "payment" in obj.action or t == "customorder":
            return "Payment"
        if t in ("product", "productvariant"):
            return "Product"
        if t in ("sitesettings", "vendor") or obj.action == "settings_update":
            return "Settings"
        if t in ("page_section", "blog_post", "hero_slide", "banner"):
            return "Content"
        return "System"


# ── Dashboard ─────────────────────────────────────────────────────────────────

class AdminDashboardView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.orders.models import Order
        from apps.users.models import User
        from apps.products.models import Product
        from apps.auctions.models import Auction

        now = timezone.now()
        thirty_days_ago = now - timedelta(days=30)

        paid_orders = Order.objects.filter(status__in=["processing", "shipped", "delivered"])
        revenue_by_currency = {
            row["currency"]: row["total"] or 0
            for row in paid_orders.values("currency").annotate(total=Sum("total"))
        }
        total_revenue_usd = revenue_by_currency.get("USD", 0)
        total_revenue_gel = revenue_by_currency.get("GEL", 0)

        return Response({
            "total_revenue": str(total_revenue_usd),
            "total_revenue_usd": str(total_revenue_usd),
            "total_revenue_gel": str(total_revenue_gel),
            "total_orders": Order.objects.count(),
            "total_users": User.objects.count(),
            "total_products": Product.objects.count(),
            "orders_last_30d": Order.objects.filter(created_at__gte=thirty_days_ago).count(),
            "new_users_last_30d": User.objects.filter(date_joined__gte=thirty_days_ago).count(),
            "active_auctions": Auction.objects.filter(ends_at__gt=now).count(),
        })


class AdminAnalyticsView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.orders.models import Order
        data = (
            Order.objects.filter(status__in=["processing", "shipped", "delivered"])
            .annotate(month=TruncMonth("created_at"))
            .values("month")
            .annotate(revenue=Sum("total"), orders=Count("id"))
            .order_by("month")
        )

        return Response([
            {
                "month": item["month"].strftime("%b %Y") if item["month"] else "",
                "revenue": str(item["revenue"] or 0),
                "orders": item["orders"],
            }
            for item in data
        ])


class AdminSuperAnalyticsView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.orders.models import Order
        from apps.auctions.models import Auction, AuctionBid
        from apps.users.models import User

        now = timezone.now()
        today = now.date()
        yesterday = today - timedelta(days=1)
        last_30_days_start = today - timedelta(days=29)
        completed_statuses = ["processing", "shipped", "delivered"]

        orders_qs = Order.objects.filter(status__in=completed_statuses)

        total_revenue = float(orders_qs.aggregate(total=Sum("total"))["total"] or 0)
        total_orders = orders_qs.count()
        total_users = User.objects.count()
        total_bids = AuctionBid.objects.count()
        total_auctions = Auction.objects.count()
        active_auctions = Auction.objects.filter(ends_at__gt=now).count()

        today_orders_qs = orders_qs.filter(created_at__date=today)
        yesterday_orders_qs = orders_qs.filter(created_at__date=yesterday)

        today_revenue = float(today_orders_qs.aggregate(total=Sum("total"))["total"] or 0)
        yesterday_revenue = float(yesterday_orders_qs.aggregate(total=Sum("total"))["total"] or 0)
        today_orders = today_orders_qs.count()
        yesterday_orders = yesterday_orders_qs.count()

        today_bids = AuctionBid.objects.filter(placed_at__date=today).count()
        yesterday_bids = AuctionBid.objects.filter(placed_at__date=yesterday).count()

        def pct_change(current, previous):
            if previous == 0:
                return 100.0 if current > 0 else 0.0
            return round(((current - previous) / previous) * 100.0, 2)

        daily_orders = (
            orders_qs.filter(created_at__date__gte=last_30_days_start)
            .annotate(day=TruncDate("created_at"))
            .values("day")
            .annotate(revenue=Sum("total"), orders=Count("id"))
            .order_by("day")
        )
        daily_bids = (
            AuctionBid.objects.filter(placed_at__date__gte=last_30_days_start)
            .annotate(day=TruncDate("placed_at"))
            .values("day")
            .annotate(bids=Count("id"), max_bid=Max("amount"))
            .order_by("day")
        )

        bids_map = {
            b["day"]: {"bids": b["bids"], "max_bid": float(b["max_bid"] or 0)}
            for b in daily_bids
        }

        by_day = []
        for i in range(30):
            day = last_30_days_start + timedelta(days=i)
            order_row = next((o for o in daily_orders if o["day"] == day), None)
            bid_row = bids_map.get(day, {"bids": 0, "max_bid": 0.0})
            by_day.append(
                {
                    "day": day.strftime("%Y-%m-%d"),
                    "orders": int(order_row["orders"]) if order_row else 0,
                    "revenue": float(order_row["revenue"] or 0) if order_row else 0.0,
                    "bids": int(bid_row["bids"]),
                    "max_bid": float(bid_row["max_bid"]),
                }
            )

        auction_bidder_rows = []
        auctions = (
            Auction.objects.select_related("product")
            .prefetch_related("bids__user")
            .order_by("-created_at")[:50]
        )
        for auction in auctions:
            bids = list(auction.bids.all().order_by("-amount", "-placed_at"))
            if not bids:
                continue
            unique_bidder_ids = set()
            bidder_list = []
            for bid in bids:
                unique_bidder_ids.add(str(bid.user_id))
                bidder_list.append(
                    {
                        "bid_id": bid.id,
                        "bidder_name": bid.user.name or bid.user.email,
                        "bidder_email": bid.user.email,
                        "amount": float(bid.amount),
                        "placed_at": bid.placed_at.isoformat(),
                    }
                )
            auction_bidder_rows.append(
                {
                    "auction_id": auction.id,
                    "title": auction.title,
                    "product_id": auction.product_id,
                    "product_title": auction.product.title if auction.product else auction.title,
                    "ends_at": auction.ends_at.isoformat(),
                    "current_bid": float(bids[0].amount),
                    "total_bids": len(bids),
                    "unique_bidders": len(unique_bidder_ids),
                    "bidders": bidder_list,
                }
            )

        auction_bidder_rows.sort(key=lambda x: (x["total_bids"], x["current_bid"]), reverse=True)

        return Response(
            {
                "totals": {
                    "revenue": total_revenue,
                    "orders": total_orders,
                    "users": total_users,
                    "bids": total_bids,
                    "auctions": total_auctions,
                    "active_auctions": active_auctions,
                },
                "today": {
                    "date": today.strftime("%Y-%m-%d"),
                    "revenue": today_revenue,
                    "orders": today_orders,
                    "bids": today_bids,
                },
                "rates": {
                    "revenue_change_pct_vs_yesterday": pct_change(today_revenue, yesterday_revenue),
                    "orders_change_pct_vs_yesterday": pct_change(today_orders, yesterday_orders),
                    "bids_change_pct_vs_yesterday": pct_change(today_bids, yesterday_bids),
                },
                "by_day": by_day,
                "auction_bidder_breakdown": auction_bidder_rows,
            }
        )


# ── Orders ────────────────────────────────────────────────────────────────────

class AdminOrderListView(AdminNoPaginationMixin, generics.ListAPIView):
    permission_classes = [IsAdminOrVendor]

    def get_serializer_class(self):
        from apps.orders.serializers import OrderSerializer
        return OrderSerializer

    def get_queryset(self):
        from apps.orders.models import Order
        qs = Order.objects.prefetch_related("items", "status_history").all()
        if not self.request.user.is_staff and hasattr(self.request.user, "vendor_profile"):
            vendor = self.request.user.vendor_profile
            qs = qs.filter(items__vendor=vendor).distinct()
        return qs


class AdminOrderUpdateView(APIView):
    permission_classes = [IsAdminOrVendor]

    def _get_order_qs(self, request):
        from apps.orders.models import Order
        qs = Order.objects.prefetch_related("items__vendor", "shipments__vendor", "status_history")
        if not request.user.is_staff:
            vendor = request.user.vendor_profile
            qs = qs.filter(items__vendor=vendor).distinct()
        return qs

    def get(self, request, pk):
        from apps.orders.serializers import OrderSerializer
        try:
            order = self._get_order_qs(request).get(pk=pk)
        except Exception:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(OrderSerializer(order).data)

    def patch(self, request, pk):
        from apps.orders.models import Order, OrderStatusHistory, OrderShipment
        from apps.orders.serializers import OrderSerializer
        try:
            order = self._get_order_qs(request).get(pk=pk)
        except Exception:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Per-shipment update: PATCH with shipment_id + tracking_code + status
        shipment_id = request.data.get("shipment_id")
        if shipment_id is not None:
            try:
                shipment = order.shipments.get(pk=int(shipment_id))
            except (OrderShipment.DoesNotExist, ValueError, TypeError):
                return Response({"detail": "Shipment not found."}, status=status.HTTP_404_NOT_FOUND)

            shipment_tracking = request.data.get("tracking_code")
            shipment_status = request.data.get("shipment_status")
            prev_shipment_status = shipment.status
            prev_shipment_tracking = shipment.tracking_code

            if shipment_tracking is not None:
                shipment.tracking_code = shipment_tracking
            if shipment_status and shipment_status in ("processing", "shipped", "delivered"):
                shipment.status = shipment_status
            should_send_shipment_email = (
                shipment.status == "shipped"
                and (
                    prev_shipment_status != "shipped"
                    or (not prev_shipment_tracking and bool(shipment.tracking_code))
                )
            )
            if should_send_shipment_email:
                if not shipment.shipped_at:
                    shipment.shipped_at = timezone.now()
                _send_shipment_email(order, shipment)

            shipment.save()

            log_action(request.user, "shipment_update", "OrderShipment", shipment.pk, {
                "order_number": order.order_number,
                "vendor": shipment.vendor.name if shipment.vendor else None,
                "tracking_code": shipment.tracking_code,
                "status": shipment.status,
            })

            # Auto-advance order status if all shipments are shipped/delivered
            all_shipped = all(
                s.status in ("shipped", "delivered") for s in order.shipments.all()
            )
            if all_shipped and order.status == "processing":
                order.status = "shipped"
                order.shipped_at = timezone.now()
                order.save(update_fields=["status", "shipped_at"])
                OrderStatusHistory.objects.create(
                    order=order, status="shipped",
                    note="All vendor shipments shipped.",
                    changed_by=request.user,
                )

            order.refresh_from_db()
            return Response(OrderSerializer(order).data)

        # Legacy single-order update
        new_status = request.data.get("status")
        note = request.data.get("note", "")
        tracking = request.data.get("tracking_code")

        prev_status = order.status
        prev_tracking = order.tracking_code

        if new_status and new_status != prev_status:
            order.status = new_status
            OrderStatusHistory.objects.create(order=order, status=new_status, note=note, changed_by=request.user)
            log_action(request.user, "order_status_change", "Order", order.pk, {
                "new_status": new_status,
                "order_number": order.order_number,
            })
            if new_status == "processing":
                log_action(request.user, "payment_received", "Order", order.pk, {
                    "order_number": order.order_number,
                    "amount": str(order.total),
                    "currency": order.currency,
                })
                try:
                    from apps.creators.services import credit_creator_for_paid_order
                    credited = credit_creator_for_paid_order(order)
                    if credited:
                        log_action(request.user, "creator_credit", "Order", order.pk, {
                            "order_number": order.order_number,
                            "promo": order.promo_code.code if order.promo_code_id else None,
                        })
                except Exception:
                    pass
            if new_status == "cancelled":
                try:
                    from apps.creators.services import clawback_creator_for_order
                    if clawback_creator_for_order(order):
                        log_action(request.user, "creator_clawback", "Order", order.pk, {
                            "order_number": order.order_number,
                        })
                except Exception:
                    pass

        if tracking is not None:
            order.tracking_code = tracking

        order.save()

        should_send_shipping_email = (
            order.status == "shipped"
            and (
                prev_status != "shipped"
                or (not prev_tracking and bool(order.tracking_code))
            )
        )
        if should_send_shipping_email:
            if not order.shipped_at:
                order.shipped_at = timezone.now()
            order.save(update_fields=["shipped_at"])
            _send_shipping_email(order)

        return Response(OrderSerializer(order).data)


def _send_shipping_email(order):
    """Send shipping confirmation — prefers HTML EmailTemplate, then vendor plain text."""
    try:
        from django.core.mail import send_mail
        from django.conf import settings as django_settings
        from apps.emails.service import send_template_email, get_template
        from apps.emails.order_context import build_order_email_context

        first_item = order.items.select_related("vendor").first()
        vendor = first_item.vendor if first_item else None
        context = build_order_email_context(order)

        # Prefer branded HTML template (vendor-specific → platform)
        if get_template("order_shipped", vendor=vendor):
            send_template_email(
                "order_shipped",
                order.shipping_email,
                context,
                vendor=vendor,
            )
            return

        customer_name = context["customer_name"]
        order_number = context["order_number"]
        tracking_info = (
            f"\n\n{context['tracking_info']}"
            if order.tracking_code else ""
        )

        if vendor and vendor.shipping_email_body:
            body = vendor.shipping_email_body
            body = body.replace("{{customer_name}}", customer_name)
            body = body.replace("{{order_number}}", order_number)
            body = body.replace("{{tracking_code}}", order.tracking_code or "")
            body = body.replace("{{tracking_info}}", tracking_info)
            subject = (vendor.shipping_email_subject or f"Your order {order_number} has shipped!")
            subject = subject.replace("{{order_number}}", order_number)
            from_email = getattr(django_settings, "EMAIL_FROM_ORDERS", None) or getattr(
                django_settings, "DEFAULT_FROM_EMAIL", "orders@koleqcia.com"
            )
        else:
            body = (
                f"Hi {customer_name},\n\n"
                f"Great news — your Koleqcia order {order_number} has been shipped!{tracking_info}\n\n"
                f"Items in your order:\n{context['items']}\n\n"
                f"Total: {context['total']}\n\n"
                f"Thank you for shopping with Koleqcia!\n"
                f"— The Koleqcia Team"
            )
            subject = f"Your order {order_number} has shipped!"
            from_email = getattr(django_settings, "EMAIL_FROM_ORDERS", None) or getattr(
                django_settings, "DEFAULT_FROM_EMAIL", "orders@koleqcia.com"
            )

        send_mail(
            subject=subject,
            message=body,
            from_email=from_email,
            recipient_list=[order.shipping_email],
            fail_silently=False,
        )
    except Exception as exc:
        logger.error("Shipping email failed for order %s to %s: %s", order.order_number, order.shipping_email, exc)


def _send_shipment_email(order, shipment):
    """Send per-vendor shipment notification when a shipment is marked shipped."""
    try:
        from django.core.mail import send_mail
        from django.conf import settings as django_settings
        from apps.emails.service import send_template_email, get_template
        from apps.emails.order_context import build_shipment_email_context

        vendor = shipment.vendor
        context = build_shipment_email_context(order, shipment)

        if get_template("order_shipped", vendor=vendor):
            send_template_email(
                "order_shipped",
                order.shipping_email,
                context,
                vendor=vendor,
            )
            return

        customer_name = context["customer_name"]
        vendor_label = vendor.name if vendor else "your vendor"
        tracking = shipment.tracking_code or ""
        tracking_info = f"\nTracking number: {tracking}" if tracking else ""
        body = (
            f"Hi {customer_name},\n\n"
            f"Items from {vendor_label} in your order {order.order_number} have shipped!{tracking_info}\n\n"
            f"Thank you for shopping with Koleqcia!\n"
            f"— The Koleqcia Team"
        )
        from_email = getattr(django_settings, "EMAIL_FROM_ORDERS", None) or getattr(
            django_settings, "DEFAULT_FROM_EMAIL", "orders@koleqcia.com"
        )
        send_mail(
            subject=f"Shipment from {vendor_label} — Order {order.order_number}",
            message=body,
            from_email=from_email,
            recipient_list=[order.shipping_email],
            fail_silently=False,
        )
    except Exception as exc:
        logger.error("Shipment email failed for order %s shipment %s: %s", order.order_number, shipment.pk, exc)


# ── Products ──────────────────────────────────────────────────────────────────

class AdminProductListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminOrVendor]

    def get_serializer_class(self):
        from apps.products.serializers import ProductDetailSerializer
        return ProductDetailSerializer

    def get_queryset(self):
        from apps.products.models import Product
        qs = Product.objects.select_related("artist", "category").prefetch_related("images", "variants__size", "variants__finish", "variants__frame", "size_variants", "categories", "processing_options")
        if not self.request.user.is_staff and hasattr(self.request.user, "vendor_profile"):
            qs = qs.filter(vendor=self.request.user.vendor_profile)
        return qs


class AdminProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminOrVendor]

    def get_serializer_class(self):
        from apps.products.serializers import ProductDetailSerializer
        return ProductDetailSerializer

    def get_queryset(self):
        from apps.products.models import Product
        qs = Product.objects.select_related("artist", "category").prefetch_related("images", "variants__size", "variants__finish", "variants__frame", "size_variants", "categories", "processing_options")
        if not self.request.user.is_staff and hasattr(self.request.user, "vendor_profile"):
            qs = qs.filter(vendor=self.request.user.vendor_profile)
        return qs

    def perform_update(self, serializer):
        instance = serializer.save()
        log_action(self.request.user, "product_update", "Product", instance.pk, {"title": instance.title})

    def perform_destroy(self, instance):
        log_action(self.request.user, "product_delete", "Product", instance.pk, {"title": instance.title})
        instance.delete()


class AdminProductStockView(APIView):
    permission_classes = [IsAdminUser]

    def patch(self, request, pk):
        from apps.products.models import ProductVariant
        try:
            variant = ProductVariant.objects.select_related("product").get(pk=pk)
        except ProductVariant.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        update_fields = []
        new_stock = request.data.get("stock")
        if new_stock is not None:
            old_stock = variant.stock
            variant.stock = int(new_stock)
            update_fields.append("stock")
            log_action(request.user, "stock_update", "ProductVariant", variant.pk, {
                "product": variant.product.title, "old_stock": old_stock, "new_stock": variant.stock
            })
        new_surcharge = request.data.get("surcharge")
        if new_surcharge is not None:
            from decimal import Decimal
            variant.surcharge = Decimal(str(new_surcharge))
            update_fields.append("surcharge")
        if not update_fields:
            return Response({"detail": "stock or surcharge is required."}, status=status.HTTP_400_BAD_REQUEST)
        variant.save(update_fields=update_fields)
        return Response({"id": variant.pk, "stock": variant.stock, "surcharge": str(variant.surcharge)})


ALLOWED_MEDIA_FOLDERS = frozenset({"blog", "hero", "categories", "auctions", "artists", "cms"})
ALLOWED_VIDEO_EXTENSIONS = frozenset({".mp4", ".webm"})
ALLOWED_VIDEO_CONTENT_TYPES = frozenset({"video/mp4", "video/webm"})
MAX_VIDEO_UPLOAD_BYTES = 80 * 1024 * 1024


class AdminMediaUploadView(APIView):
    """Generic admin media upload for CMS and catalog assets."""
    permission_classes = [IsAdminOrVendor]

    def post(self, request):
        import os
        import uuid as uuid_lib
        from django.conf import settings as django_settings
        from apps.core.uploads import validate_image_upload, safe_image_extension

        uploaded_file = request.FILES.get("file")
        folder = (request.data.get("folder") or "cms").strip().lower()
        # Security fix #14: path traversal prevention
        if folder not in ALLOWED_MEDIA_FOLDERS or os.path.basename(folder) != folder:
            return Response(
                {"detail": f"folder must be one of: {', '.join(sorted(ALLOWED_MEDIA_FOLDERS))}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        name = getattr(uploaded_file, "name", "") or ""
        raw_ext = os.path.splitext(name)[1].lower()
        content_type = (getattr(uploaded_file, "content_type", "") or "").lower()
        is_video = content_type.startswith("video/") or raw_ext in ALLOWED_VIDEO_EXTENSIONS

        if is_video:
            if raw_ext not in ALLOWED_VIDEO_EXTENSIONS or content_type not in ALLOWED_VIDEO_CONTENT_TYPES:
                return Response(
                    {"detail": "Unsupported video format. Upload an H.264 MP4 or WebM file."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if getattr(uploaded_file, "size", 0) > MAX_VIDEO_UPLOAD_BYTES:
                return Response({"detail": "Video is too large. Maximum upload size is 80 MB."}, status=status.HTTP_400_BAD_REQUEST)
            ext = raw_ext
        else:
            error = validate_image_upload(uploaded_file)
            if error:
                return error
            ext = safe_image_extension(uploaded_file)
        filename = f"{uuid_lib.uuid4().hex}{ext}"
        save_dir = os.path.join(django_settings.MEDIA_ROOT, folder)
        os.makedirs(save_dir, exist_ok=True)
        save_path = os.path.join(save_dir, filename)
        with open(save_path, "wb") as f:
            for chunk in uploaded_file.chunks():
                f.write(chunk)

        url = request.build_absolute_uri(f"{django_settings.MEDIA_URL}{folder}/{filename}")
        return Response({"url": url, "folder": folder}, status=status.HTTP_201_CREATED)


class AdminProductMediaView(APIView):
    """Upload a video (or image file) to a product's media gallery."""
    permission_classes = [IsAdminOrVendor]

    def post(self, request):
        from apps.products.models import Product, ProductImage
        from apps.products.serializers import ProductImageSerializer
        product_id = request.data.get("product_id")
        uploaded_file = request.FILES.get("file")
        media_type = request.data.get("media_type", "image")

        if not product_id or not uploaded_file:
            return Response({"detail": "product_id and file are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            return Response({"detail": "Product not found."}, status=status.HTTP_404_NOT_FOUND)

        if not request.user.is_staff and hasattr(request.user, "vendor_profile"):
            if product.vendor_id != request.user.vendor_profile.id:
                return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        max_order = product.images.count()

        img = ProductImage.objects.create(
            product=product,
            video_file=uploaded_file,
            media_type=media_type,
            order=max_order,
        )

        return Response(ProductImageSerializer(img, context={"request": request}).data, status=status.HTTP_201_CREATED)

    def delete(self, request, image_id):
        from apps.products.models import ProductImage
        try:
            img = ProductImage.objects.get(pk=image_id)
            if img.video_file:
                img.video_file.delete(save=False)
            img.delete()
        except ProductImage.DoesNotExist:
            pass
        return Response(status=status.HTTP_204_NO_CONTENT)


class AdminProductMediaReorderView(APIView):
    """PATCH /admin/products/media/reorder/ — accept [{id, order}] and bulk-update ProductImage.order."""
    permission_classes = [IsAdminOrVendor]

    def patch(self, request):
        from apps.products.models import ProductImage
        items = request.data if isinstance(request.data, list) else request.data.get("items", [])
        if not items:
            return Response({"detail": "items list required."}, status=status.HTTP_400_BAD_REQUEST)
        ids = [item["id"] for item in items if "id" in item]
        images = {img.pk: img for img in ProductImage.objects.filter(pk__in=ids)}
        updated = []
        for item in items:
            img = images.get(item.get("id"))
            if img is not None and "order" in item:
                img.order = item["order"]
                updated.append(img)
        ProductImage.objects.bulk_update(updated, ["order"])
        return Response({"updated": len(updated)})


# ── Categories ────────────────────────────────────────────────────────────────

class AdminCategoryListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminOrVendor]

    def get_serializer_class(self):
        from apps.products.serializers import CategorySerializer
        return CategorySerializer

    def get_queryset(self):
        from apps.products.models import Category
        return Category.objects.all()


class AdminCategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminOrVendor]

    def get_serializer_class(self):
        from apps.products.serializers import CategorySerializer
        return CategorySerializer

    def get_queryset(self):
        from apps.products.models import Category
        return Category.objects.all()


# ── Artists ───────────────────────────────────────────────────────────────────

class AdminArtistListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.products.serializers import ArtistSerializer
        return ArtistSerializer

    def get_queryset(self):
        from apps.products.models import Artist
        return Artist.objects.all()


class AdminArtistDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.products.serializers import ArtistSerializer
        return ArtistSerializer

    def get_queryset(self):
        from apps.products.models import Artist
        return Artist.objects.all()


# ── Users ─────────────────────────────────────────────────────────────────────

class AdminUserListView(AdminNoPaginationMixin, generics.ListAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.users.serializers import UserSerializer
        return UserSerializer

    def get_queryset(self):
        from apps.users.models import User
        return User.objects.all().order_by("-date_joined")


class AdminUserToggleView(APIView):
    permission_classes = [IsAdminUser]

    def patch(self, request, pk):
        from apps.users.models import User
        from apps.users.serializers import UserSerializer
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        user.is_active = not user.is_active
        user.save(update_fields=["is_active"])
        action = "user_activate" if user.is_active else "user_deactivate"
        log_action(request.user, action, "User", user.pk, {"email": user.email})
        return Response({"id": str(user.pk), "is_active": user.is_active})


# ── Promo Codes ───────────────────────────────────────────────────────────────

class AdminPromoListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.promo.serializers import PromoCodeSerializer
        return PromoCodeSerializer

    def get_queryset(self):
        from apps.promo.models import PromoCode
        return PromoCode.objects.all().order_by("-created_at")

    def perform_create(self, serializer):
        instance = serializer.save()
        log_action(self.request.user, "promo_create", "PromoCode", instance.pk, {"code": instance.code})


class AdminPromoDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.promo.serializers import PromoCodeSerializer
        return PromoCodeSerializer

    def get_queryset(self):
        from apps.promo.models import PromoCode
        return PromoCode.objects.all()

    def perform_update(self, serializer):
        instance = serializer.save()
        log_action(self.request.user, "promo_update", "PromoCode", instance.pk, {"code": instance.code})

    def perform_destroy(self, instance):
        log_action(self.request.user, "promo_delete", "PromoCode", instance.pk, {"code": instance.code})
        instance.delete()


# ── Reviews ───────────────────────────────────────────────────────────────────

class AdminHomepageReviewListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.cms.serializers import HomepageReviewSerializer
        return HomepageReviewSerializer

    def get_queryset(self):
        from apps.cms.models import HomepageReview
        return HomepageReview.objects.all()


class AdminHomepageReviewDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.cms.serializers import HomepageReviewSerializer
        return HomepageReviewSerializer

    def get_queryset(self):
        from apps.cms.models import HomepageReview
        return HomepageReview.objects.all()


class AdminCommunitySocialLinkListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.cms.serializers import CommunitySocialLinkSerializer
        return CommunitySocialLinkSerializer

    def get_queryset(self):
        from apps.cms.models import CommunitySocialLink
        from apps.cms.defaults import ensure_global_homepage_defaults
        ensure_global_homepage_defaults()
        return CommunitySocialLink.objects.all()


class AdminCommunitySocialLinkDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.cms.serializers import CommunitySocialLinkSerializer
        return CommunitySocialLinkSerializer

    def get_queryset(self):
        from apps.cms.models import CommunitySocialLink
        return CommunitySocialLink.objects.all()


class AdminTrustBarItemListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]
    def get_serializer_class(self):
        from apps.cms.serializers import TrustBarItemSerializer
        return TrustBarItemSerializer
    def get_queryset(self):
        from apps.cms.models import TrustBarItem
        from apps.cms.defaults import ensure_global_homepage_defaults
        ensure_global_homepage_defaults()
        return TrustBarItem.objects.all()


class AdminTrustBarItemDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]
    def get_serializer_class(self):
        from apps.cms.serializers import TrustBarItemSerializer
        return TrustBarItemSerializer
    def get_queryset(self):
        from apps.cms.models import TrustBarItem
        return TrustBarItem.objects.all()


class AdminFandomBrandListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]
    def get_serializer_class(self):
        from apps.cms.serializers import FandomBrandSerializer
        return FandomBrandSerializer
    def get_queryset(self):
        from apps.cms.models import FandomBrand
        from apps.cms.defaults import ensure_global_homepage_defaults
        ensure_global_homepage_defaults()
        return FandomBrand.objects.all()


class AdminFandomBrandDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]
    def get_serializer_class(self):
        from apps.cms.serializers import FandomBrandSerializer
        return FandomBrandSerializer
    def get_queryset(self):
        from apps.cms.models import FandomBrand
        return FandomBrand.objects.all()


class AdminFAQListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.cms.serializers import FAQSerializer
        return FAQSerializer

    def get_queryset(self):
        from apps.cms.models import FAQ
        qs = FAQ.objects.all()
        category = self.request.query_params.get("category")
        if category:
            return qs.filter(category=category)
        return qs


class AdminFAQDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.cms.serializers import FAQSerializer
        return FAQSerializer

    def get_queryset(self):
        from apps.cms.models import FAQ
        return FAQ.objects.all()


class AdminReviewListView(AdminNoPaginationMixin, generics.ListAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.products.serializers import ReviewSerializer
        return ReviewSerializer

    def get_queryset(self):
        from apps.products.models import Review
        return Review.objects.select_related("user", "product").all()


class AdminReviewApproveView(APIView):
    permission_classes = [IsAdminUser]

    def patch(self, request, pk):
        from apps.products.models import Review
        from apps.products.serializers import ReviewSerializer
        try:
            review = Review.objects.get(pk=pk)
        except Review.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        review.approved = not review.approved
        review.save(update_fields=["approved"])
        return Response(ReviewSerializer(review).data)


# ── Auctions ──────────────────────────────────────────────────────────────────

class AdminAuctionListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.auctions.serializers import AuctionSerializer, AuctionWriteSerializer
        if self.request.method == "POST":
            return AuctionWriteSerializer
        return AuctionSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["include_all_bids"] = True
        return ctx

    def get_queryset(self):
        from apps.auctions.models import Auction
        return Auction.objects.prefetch_related("bids__user").select_related("product", "vendor", "winner").all()


class AdminAuctionDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.auctions.serializers import AuctionSerializer, AuctionWriteSerializer
        if self.request.method in ("PUT", "PATCH"):
            return AuctionWriteSerializer
        return AuctionSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["include_all_bids"] = True
        return ctx

    def get_queryset(self):
        from apps.auctions.models import Auction
        return Auction.objects.prefetch_related("bids__user").select_related("product", "vendor", "winner").all()


# ── CMS (Hero / Banners) ──────────────────────────────────────────────────────

class AdminHeroListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.cms.serializers import HeroSlideSerializer
        return HeroSlideSerializer

    def get_queryset(self):
        from apps.cms.models import HeroSlide
        return HeroSlide.objects.all()


class AdminHeroDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.cms.serializers import HeroSlideSerializer
        return HeroSlideSerializer

    def get_queryset(self):
        from apps.cms.models import HeroSlide
        return HeroSlide.objects.all()


class AdminBannerListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.cms.serializers import BannerSerializer
        return BannerSerializer

    def get_queryset(self):
        from apps.cms.models import Banner
        return Banner.objects.all()


class AdminBannerDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.cms.serializers import BannerSerializer
        return BannerSerializer

    def get_queryset(self):
        from apps.cms.models import Banner
        return Banner.objects.all()


class AdminAnnouncementBarView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.cms.models import AnnouncementBar
        from apps.cms.serializers import AnnouncementBarSerializer
        bar, _ = AnnouncementBar.objects.get_or_create(
            pk=1,
            defaults={
                "messages": [
                    "FREE SHIPPING on orders over $49 — use code FREESHIP",
                    "LIMITED EDITIONS: New drops every Friday at noon",
                    "EARN XP with every purchase — unlock exclusive badges",
                ],
                "is_active": True,
            },
        )
        return Response(AnnouncementBarSerializer(bar).data)

    def patch(self, request):
        from apps.cms.models import AnnouncementBar
        from apps.cms.serializers import AnnouncementBarSerializer
        bar, _ = AnnouncementBar.objects.get_or_create(pk=1)
        ser = AnnouncementBarSerializer(bar, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)


# ── Newsletter ────────────────────────────────────────────────────────────────

class AdminNewsletterListView(AdminNoPaginationMixin, generics.ListAPIView):
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        from apps.newsletter.models import NewsletterSubscriber
        return NewsletterSubscriber.objects.all().order_by("-subscribed_at")

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        data = [{"email": s.email, "subscribed_at": s.subscribed_at} for s in qs]
        return Response({"count": len(data), "results": data})


# ── Inbox (Admin) ─────────────────────────────────────────────────────────────

class AdminConversationListView(AdminNoPaginationMixin, generics.ListAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.messaging.serializers import ConversationSerializer
        return ConversationSerializer

    def get_queryset(self):
        from apps.messaging.models import Conversation
        return Conversation.objects.prefetch_related("messages").all()


# ── Settings ──────────────────────────────────────────────────────────────────

class AdminSettingsView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.cms.models import SiteSettings
        settings = SiteSettings.objects.all()
        return Response({s.key: s.value for s in settings})

    def patch(self, request):
        from apps.cms.models import SiteSettings
        keys = []
        for key, value in request.data.items():
            SiteSettings.objects.update_or_create(key=key, defaults={"value": str(value)})
            keys.append(key)
        log_action(request.user, "settings_update", "SiteSettings", "global", {"keys": keys})
        return Response({"detail": "Settings updated."})


# ── Gamification ──────────────────────────────────────────────────────────────

class AdminBadgeListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.gamification.serializers import BadgeSerializer, BadgeWriteSerializer
        if self.request.method == "POST":
            return BadgeWriteSerializer
        return BadgeSerializer

    def get_queryset(self):
        from apps.gamification.models import Badge
        return Badge.objects.select_related("prize_promo").all()

    def create(self, request, *args, **kwargs):
        from apps.gamification.serializers import BadgeSerializer, BadgeWriteSerializer
        write_ser = BadgeWriteSerializer(data=request.data)
        write_ser.is_valid(raise_exception=True)
        badge = write_ser.save()
        badge = self.get_queryset().get(pk=badge.pk)
        return Response(BadgeSerializer(badge).data, status=status.HTTP_201_CREATED)


class AdminBadgeDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.gamification.serializers import BadgeSerializer, BadgeWriteSerializer
        if self.request.method in ("PUT", "PATCH"):
            return BadgeWriteSerializer
        return BadgeSerializer

    def get_queryset(self):
        from apps.gamification.models import Badge
        return Badge.objects.select_related("prize_promo").all()

    def update(self, request, *args, **kwargs):
        from apps.gamification.serializers import BadgeSerializer, BadgeWriteSerializer
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        write_ser = BadgeWriteSerializer(instance, data=request.data, partial=partial)
        write_ser.is_valid(raise_exception=True)
        write_ser.save()
        instance = self.get_queryset().get(pk=instance.pk)
        return Response(BadgeSerializer(instance).data)


class AdminXPRuleListView(AdminNoPaginationMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.gamification.serializers import XPRuleSerializer
        return XPRuleSerializer

    def get_queryset(self):
        from apps.gamification.models import XPRule
        return XPRule.objects.all().order_by("action_key")


class AdminXPRuleDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]

    def get_serializer_class(self):
        from apps.gamification.serializers import XPRuleSerializer
        return XPRuleSerializer

    def get_queryset(self):
        from apps.gamification.models import XPRule
        return XPRule.objects.all()


# ── Audit Log ─────────────────────────────────────────────────────────────────

class AuditLogPagination(PageNumberPagination):
    page_size = 100
    page_size_query_param = "page_size"
    max_page_size = 500


class AdminAuditLogView(generics.ListAPIView):
    serializer_class = AuditLogSerializer
    permission_classes = [IsAdminUser]
    pagination_class = AuditLogPagination

    def get_queryset(self):
        qs = AuditLog.objects.select_related("admin_user").all()
        action = self.request.query_params.get("action")
        target_type = self.request.query_params.get("target_type")
        category = self.request.query_params.get("category")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        search = self.request.query_params.get("search", "").strip()

        if action:
            qs = qs.filter(action=action)
        if target_type:
            qs = qs.filter(target_type__iexact=target_type)
        if date_from:
            qs = qs.filter(timestamp__date__gte=date_from)
        if date_to:
            qs = qs.filter(timestamp__date__lte=date_to)
        if search:
            qs = qs.filter(
                Q(target_id__icontains=search)
                | Q(detail__icontains=search)
                | Q(admin_user__email__icontains=search)
            )
        if category:
            cat = category.lower()
            if cat == "order":
                qs = qs.filter(Q(target_type__iexact="Order") | Q(action__startswith="order"))
            elif cat == "payment":
                qs = qs.filter(Q(action__icontains="payment") | Q(target_type__iexact="CustomOrder"))
            elif cat == "product":
                qs = qs.filter(target_type__in=["Product", "ProductVariant"])
            elif cat == "settings":
                qs = qs.filter(Q(target_type__in=["SiteSettings", "Vendor"]) | Q(action="settings_update"))
            elif cat == "content":
                qs = qs.filter(target_type__in=["page_section", "blog_post", "HeroSlide", "Banner"])
        return qs


class AdminBlogPostSerializer(serializers.ModelSerializer):
    class Meta:
        from apps.blog.models import BlogPost
        model = BlogPost
        fields = (
            "id",
            "title",
            "title_ka",
            "title_ru",
            "slug",
            "excerpt",
            "excerpt_ka",
            "excerpt_ru",
            "content",
            "content_ka",
            "content_ru",
            "content_blocks",
            "cover_image_url",
            "is_published",
            "published_at",
            "created_at",
            "updated_at",
        )
        read_only_fields = ("id", "slug", "created_at", "updated_at")


class AdminBlogPostListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.blog.models import BlogPost
        posts = BlogPost.objects.all().order_by("-created_at")
        return Response(AdminBlogPostSerializer(posts, many=True).data)

    def post(self, request):
        payload = request.data.copy()
        if payload.get("is_published") and not payload.get("published_at"):
            payload["published_at"] = timezone.now()
        ser = AdminBlogPostSerializer(data=payload)
        ser.is_valid(raise_exception=True)
        post = ser.save()
        log_action(request.user, "create", "blog_post", post.id, {"title": post.title})
        return Response(AdminBlogPostSerializer(post).data, status=status.HTTP_201_CREATED)


class AdminBlogPostDetailView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request, pk):
        from apps.blog.models import BlogPost
        try:
            post = BlogPost.objects.get(pk=pk)
        except BlogPost.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(AdminBlogPostSerializer(post).data)

    def patch(self, request, pk):
        from apps.blog.models import BlogPost
        try:
            post = BlogPost.objects.get(pk=pk)
        except BlogPost.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        payload = request.data.copy()
        if payload.get("is_published") and not post.published_at and not payload.get("published_at"):
            payload["published_at"] = timezone.now()
        ser = AdminBlogPostSerializer(post, data=payload, partial=True)
        ser.is_valid(raise_exception=True)
        updated = ser.save()
        log_action(request.user, "update", "blog_post", updated.id, {"title": updated.title})
        return Response(AdminBlogPostSerializer(updated).data)

    def delete(self, request, pk):
        from apps.blog.models import BlogPost
        try:
            post = BlogPost.objects.get(pk=pk)
        except BlogPost.DoesNotExist:
            return Response(status=status.HTTP_204_NO_CONTENT)
        title = post.title
        post.delete()
        log_action(request.user, "delete", "blog_post", pk, {"title": title})
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Delivery options ──────────────────────────────────────────────────────────

class AdminDeliveryOptionListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.orders.models import DeliveryOption
        from apps.orders.serializers import DeliveryOptionSerializer
        opts = DeliveryOption.objects.all()
        return Response(DeliveryOptionSerializer(opts, many=True).data)

    def post(self, request):
        from apps.orders.models import DeliveryOption
        from apps.orders.serializers import DeliveryOptionSerializer
        ser = DeliveryOptionSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        opt = ser.save()
        return Response(DeliveryOptionSerializer(opt).data, status=status.HTTP_201_CREATED)


class AdminDeliveryOptionDetailView(APIView):
    permission_classes = [IsAdminUser]

    def patch(self, request, pk):
        from apps.orders.models import DeliveryOption
        from apps.orders.serializers import DeliveryOptionSerializer
        try:
            opt = DeliveryOption.objects.get(pk=pk)
        except DeliveryOption.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        ser = DeliveryOptionSerializer(opt, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(DeliveryOptionSerializer(opt).data)

    def delete(self, request, pk):
        from apps.orders.models import DeliveryOption
        try:
            DeliveryOption.objects.get(pk=pk).delete()
        except DeliveryOption.DoesNotExist:
            pass
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Processing options ─────────────────────────────────────────────────────────

class AdminProcessingOptionListView(APIView):
    permission_classes = [IsAdminOrVendor]

    def _vendor_from_request(self, request):
        from apps.vendors.models import Vendor
        slug = request.query_params.get("vendor") or request.data.get("vendor_slug")
        if slug:
            return Vendor.objects.filter(slug=slug).first()
        # If the user is a vendor admin, auto-resolve their vendor
        if hasattr(request.user, "vendor_profile"):
            return request.user.vendor_profile
        return None

    def get(self, request):
        from apps.orders.models import ProcessingOption
        from apps.orders.serializers import ProcessingOptionSerializer
        vendor = self._vendor_from_request(request)
        opts = ProcessingOption.objects.select_related("vendor").all()
        if vendor:
            opts = opts.filter(vendor=vendor)
        elif not request.user.is_staff:
            # Non-staff without a vendor — return empty
            opts = opts.none()
        return Response(ProcessingOptionSerializer(opts, many=True).data)

    def post(self, request):
        from apps.orders.models import ProcessingOption
        from apps.orders.serializers import ProcessingOptionSerializer
        from apps.vendors.models import Vendor
        from django.utils.text import slugify
        import uuid
        data = request.data.copy()
        vendor = self._vendor_from_request(request)
        slug = data.pop("vendor_slug", None)
        if not vendor and slug:
            vendor = Vendor.objects.filter(slug=slug).first()
        # Always set vendor (None if not found)
        data["vendor"] = vendor.id if vendor else None
        # Auto-generate slug from label if not provided
        if "slug" not in data or not data["slug"]:
            base = slugify(data.get("label", "option"))
            data["slug"] = base or str(uuid.uuid4())[:8]
        ser = ProcessingOptionSerializer(data=data)
        ser.is_valid(raise_exception=True)
        opt = ser.save()
        return Response(ProcessingOptionSerializer(opt).data, status=status.HTTP_201_CREATED)


class AdminProcessingOptionDetailView(APIView):
    permission_classes = [IsAdminOrVendor]

    def _get_option(self, request, pk):
        """Security fix #13: enforce vendor isolation on processing options."""
        from apps.orders.models import ProcessingOption
        try:
            opt = ProcessingOption.objects.select_related("vendor").get(pk=pk)
        except ProcessingOption.DoesNotExist:
            return None
        if not request.user.is_staff and hasattr(request.user, "vendor_profile"):
            if opt.vendor_id != request.user.vendor_profile.id:
                return None
        return opt

    def patch(self, request, pk):
        from apps.orders.serializers import ProcessingOptionSerializer
        opt = self._get_option(request, pk)
        if not opt:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        ser = ProcessingOptionSerializer(opt, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ProcessingOptionSerializer(opt).data)

    def delete(self, request, pk):
        opt = self._get_option(request, pk)
        if not opt:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        opt.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Size variants ──────────────────────────────────────────────────────────────

class AdminSizeVariantView(APIView):
    permission_classes = [IsAdminOrVendor]

    def post(self, request):
        from apps.products.models import Product, SizeVariant
        from apps.products.serializers import SizeVariantSerializer
        product_id = request.data.get("product_id")
        if not product_id:
            return Response({"detail": "product_id required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            return Response({"detail": "Product not found."}, status=status.HTTP_404_NOT_FOUND)
        if not request.user.is_staff and hasattr(request.user, "vendor_profile"):
            if product.vendor_id != request.user.vendor_profile.id:
                return Response({"detail": "Product not found."}, status=status.HTTP_404_NOT_FOUND)
        ser = SizeVariantSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        image_ids = ser.validated_data.pop("images", None)
        sv = SizeVariant.objects.create(product=product, **ser.validated_data)
        if image_ids is not None:
            sv.images.set(image_ids)
        return Response(SizeVariantSerializer(sv).data, status=status.HTTP_201_CREATED)

    def patch(self, request, sv_id):
        from apps.products.models import SizeVariant
        from apps.products.serializers import SizeVariantSerializer
        try:
            sv = SizeVariant.objects.get(pk=sv_id)
        except SizeVariant.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        ser = SizeVariantSerializer(sv, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        image_ids = ser.validated_data.pop("images", None)
        ser.save()
        if image_ids is not None:
            sv.images.set(image_ids)
        return Response(SizeVariantSerializer(sv).data)

    def delete(self, request, sv_id):
        from apps.products.models import SizeVariant
        try:
            SizeVariant.objects.get(pk=sv_id).delete()
        except SizeVariant.DoesNotExist:
            pass
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Product export / import ────────────────────────────────────────────────────

class AdminProductExportView(APIView):
    permission_classes = [IsAdminOrVendor]

    def get(self, request):
        try:
            import openpyxl
        except ImportError:
            return Response({"detail": "openpyxl not installed."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        from apps.products.models import Product, SizeVariant
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        headers = [
            "id", "title", "title_ka", "title_ru", "description", "description_ka", "description_ru", "material", "material_ka", "material_ru",
            "base_price_usd", "price_gel", "price_eur", "price_gbp",
            "categories", "allow_custom_size",
            "status", "is_limited", "is_sale", "is_new", "is_exclusive", "is_featured", "is_ready_to_ship",
            "tags", "tags_ka", "tags_ru", "product_details", "product_details_ka", "product_details_ru", "vendor_slug",
            "image_url_1", "image_url_2", "image_url_3",
            "size_1_label", "size_1_label_ka", "size_1_label_ru", "size_1_sku", "size_1_stock", "size_1_ready_to_ship", "size_1_price_usd", "size_1_price_gel", "size_1_sale_usd", "size_1_sale_gel",
            "size_2_label", "size_2_label_ka", "size_2_label_ru", "size_2_sku", "size_2_stock", "size_2_ready_to_ship", "size_2_price_usd", "size_2_price_gel", "size_2_sale_usd", "size_2_sale_gel",
            "size_3_label", "size_3_label_ka", "size_3_label_ru", "size_3_sku", "size_3_stock", "size_3_ready_to_ship", "size_3_price_usd", "size_3_price_gel", "size_3_sale_usd", "size_3_sale_gel",
            "size_4_label", "size_4_label_ka", "size_4_label_ru", "size_4_sku", "size_4_stock", "size_4_ready_to_ship", "size_4_price_usd", "size_4_price_gel", "size_4_sale_usd", "size_4_sale_gel",
            "size_5_label", "size_5_label_ka", "size_5_label_ru", "size_5_sku", "size_5_stock", "size_5_ready_to_ship", "size_5_price_usd", "size_5_price_gel", "size_5_sale_usd", "size_5_sale_gel",
        ]
        ws.append(headers)

        products = Product.objects.prefetch_related("images", "size_variants", "categories", "vendor").filter(status="active")
        if not request.user.is_staff and hasattr(request.user, "vendor_profile"):
            products = products.filter(vendor=request.user.vendor_profile)
        for p in products:
            images = list(p.images.values_list("url", flat=True))[:3]
            while len(images) < 3:
                images.append("")
            svs = list(
                p.size_variants.filter(is_active=True).values_list(
                    "label", "label_ka", "label_ru", "sku", "stock", "is_ready_to_ship", "price_usd", "price_gel", "sale_price_usd", "sale_price_gel"
                )
            )[:5]
            while len(svs) < 5:
                svs.append(("", "", "", "", "", "", "", "", "", ""))
            flat_svs = []
            for lbl, lbl_ka, lbl_ru, sku, stock, ready, pr_usd, pr_gel, sale_usd, sale_gel in svs:
                flat_svs.extend([
                    lbl, lbl_ka or "", lbl_ru or "", sku or "", stock if stock is not None else "", "yes" if ready else "no",
                    str(pr_usd) if pr_usd is not None else "",
                    str(pr_gel) if pr_gel is not None else "",
                    str(sale_usd) if sale_usd is not None else "",
                    str(sale_gel) if sale_gel is not None else "",
                ])
            rp = p.regional_prices or {}
            row = [
                p.id, p.title, getattr(p, "title_ka", "") or "", getattr(p, "title_ru", "") or "",
                p.description, getattr(p, "description_ka", "") or "", getattr(p, "description_ru", "") or "",
                p.material, getattr(p, "material_ka", "") or "", getattr(p, "material_ru", "") or "",
                str(p.base_price),
                str(rp.get("GEL", {}).get("price", "") or ""),
                str(rp.get("EUR", {}).get("price", "") or ""),
                str(rp.get("GBP", {}).get("price", "") or ""),
                ",".join(p.categories.values_list("slug", flat=True)),
                "yes" if p.allow_custom_size else "no",
                p.status,
                "yes" if p.is_limited else "no",
                "yes" if p.is_sale else "no",
                "yes" if p.is_new else "no",
                "yes" if p.is_exclusive else "no",
                "yes" if p.is_featured else "no",
                "yes" if p.is_ready_to_ship else "no",
                ",".join(p.tags or []),
                ",".join(getattr(p, "tags_ka", None) or []),
                ",".join(getattr(p, "tags_ru", None) or []),
                "|".join(p.product_details or []),
                "|".join(getattr(p, "product_details_ka", None) or []),
                "|".join(getattr(p, "product_details_ru", None) or []),
                p.vendor.slug if p.vendor else "",
            ] + images + flat_svs
            ws.append(row)

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="products_export.xlsx"'
        return response


class AdminProductImportView(APIView):
    permission_classes = [IsAdminOrVendor]

    def get(self, request):
        """Download blank template."""
        try:
            import openpyxl
        except ImportError:
            return Response({"detail": "openpyxl not installed."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        headers = [
            "id", "title", "title_ka", "title_ru", "description", "description_ka", "description_ru", "material", "material_ka", "material_ru",
            "base_price_usd", "price_gel", "price_eur", "price_gbp",
            "categories", "allow_custom_size",
            "status", "is_limited", "is_sale", "is_new", "is_exclusive", "is_featured", "is_ready_to_ship",
            "tags", "tags_ka", "tags_ru", "product_details", "product_details_ka", "product_details_ru", "vendor_slug",
            "image_url_1", "image_url_2", "image_url_3",
            "size_1_label", "size_1_label_ka", "size_1_label_ru", "size_1_sku", "size_1_stock", "size_1_ready_to_ship", "size_1_price_usd", "size_1_price_gel", "size_1_sale_usd", "size_1_sale_gel",
            "size_2_label", "size_2_label_ka", "size_2_label_ru", "size_2_sku", "size_2_stock", "size_2_ready_to_ship", "size_2_price_usd", "size_2_price_gel", "size_2_sale_usd", "size_2_sale_gel",
            "size_3_label", "size_3_label_ka", "size_3_label_ru", "size_3_sku", "size_3_stock", "size_3_ready_to_ship", "size_3_price_usd", "size_3_price_gel", "size_3_sale_usd", "size_3_sale_gel",
            "size_4_label", "size_4_label_ka", "size_4_label_ru", "size_4_sku", "size_4_stock", "size_4_ready_to_ship", "size_4_price_usd", "size_4_price_gel", "size_4_sale_usd", "size_4_sale_gel",
            "size_5_label", "size_5_label_ka", "size_5_label_ru", "size_5_sku", "size_5_stock", "size_5_ready_to_ship", "size_5_price_usd", "size_5_price_gel", "size_5_sale_usd", "size_5_sale_gel",
        ]
        ws.append(headers)
        ws.append([
            "", "Example Product", "", "", "A beautiful piece.", "", "", "Metal", "", "",
            "49.99", "135.00", "46.00", "39.50",
            "figures", "no",
            "active", "no", "yes", "yes", "no", "no", "yes",
            "art,modern", "", "", "Premium print|Magnetic mount included", "", "", "example-vendor",
            "https://example.com/img1.jpg", "", "",
            "S", "", "", "", "10", "yes", "39.99", "108.00", "34.99", "95.00",
            "M", "", "", "", "", "no", "49.99", "135.00", "44.99", "120.00",
            "L", "", "", "", "", "no", "59.99", "162.00", "", "",
            "", "", "", "", "", "", "", "", "", "",
            "", "", "", "", "", "", "", "", "", "",
        ])
        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="products_template.xlsx"'
        return response

    def post(self, request):
        """Import products from uploaded xlsx."""
        try:
            import openpyxl
        except ImportError:
            return Response({"detail": "openpyxl not installed."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        from apps.products.models import Product, ProductImage, SizeVariant, Category
        from apps.vendors.models import Vendor

        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
        except Exception as e:
            return Response({"detail": f"Invalid xlsx: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Empty file."}, status=status.HTTP_400_BAD_REQUEST)

        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        data_rows = rows[1:]
        created_count = 0
        updated_count = 0
        errors = []

        def cell(row, name):
            try:
                idx = headers.index(name)
                return row[idx] if idx < len(row) else None
            except ValueError:
                return None

        def yn(val):
            return str(val).strip().lower() in ("yes", "true", "1")

        def csv_list(val):
            return [t.strip() for t in str(val or "").split(",") if t.strip()]

        def pipe_list(val):
            return [t.strip() for t in str(val or "").split("|") if t.strip()]

        for i, row in enumerate(data_rows, start=2):
            title = cell(row, "title")
            if not title:
                continue
            try:
                from decimal import Decimal as D
                base_price = D(str(cell(row, "base_price_usd") or "0"))
                regional_prices = {}
                for cur, col in [("GEL", "price_gel"), ("EUR", "price_eur"), ("GBP", "price_gbp")]:
                    v = cell(row, col)
                    if v:
                        regional_prices[cur] = {"price": str(v)}

                tags = csv_list(cell(row, "tags"))

                vendor_slug = cell(row, "vendor_slug")
                vendor = Vendor.objects.filter(slug=vendor_slug).first() if vendor_slug else None
                if not vendor and not request.user.is_staff and hasattr(request.user, "vendor_profile"):
                    vendor = request.user.vendor_profile

                cat_raw = cell(row, "categories") or ""
                cat_slugs = [s.strip() for s in str(cat_raw).split(",") if str(s).strip()]
                cats = list(Category.objects.filter(slug__in=cat_slugs))
                primary_cat = cats[0] if cats else None

                product_id = cell(row, "id")
                product = None
                if product_id:
                    qs = Product.objects.filter(pk=product_id)
                    if not request.user.is_staff and hasattr(request.user, "vendor_profile"):
                        qs = qs.filter(vendor=request.user.vendor_profile)
                    product = qs.first()

                payload = {
                    "title": str(title),
                    "title_ka": str(cell(row, "title_ka") or ""),
                    "title_ru": str(cell(row, "title_ru") or ""),
                    "description": str(cell(row, "description") or ""),
                    "description_ka": str(cell(row, "description_ka") or ""),
                    "description_ru": str(cell(row, "description_ru") or ""),
                    "material": str(cell(row, "material") or ""),
                    "material_ka": str(cell(row, "material_ka") or ""),
                    "material_ru": str(cell(row, "material_ru") or ""),
                    "base_price": base_price,
                    "regional_prices": regional_prices,
                    "allow_custom_size": yn(cell(row, "allow_custom_size")),
                    "status": str(cell(row, "status") or "active"),
                    "is_limited": yn(cell(row, "is_limited")),
                    "is_sale": yn(cell(row, "is_sale")),
                    "is_new": yn(cell(row, "is_new")),
                    "is_exclusive": yn(cell(row, "is_exclusive")),
                    "is_featured": yn(cell(row, "is_featured")),
                    "is_ready_to_ship": yn(cell(row, "is_ready_to_ship")),
                    "tags": tags,
                    "tags_ka": csv_list(cell(row, "tags_ka")),
                    "tags_ru": csv_list(cell(row, "tags_ru")),
                    "product_details": pipe_list(cell(row, "product_details")),
                    "product_details_ka": pipe_list(cell(row, "product_details_ka")),
                    "product_details_ru": pipe_list(cell(row, "product_details_ru")),
                    "category": primary_cat,
                    "vendor": vendor,
                }
                if product:
                    for field, value in payload.items():
                        setattr(product, field, value)
                    product.save()
                    updated_count += 1
                    product.images.all().delete()
                    product.size_variants.all().delete()
                else:
                    product = Product.objects.create(**payload)
                    created_count += 1
                if cats:
                    product.categories.set(cats)

                for n in range(1, 4):
                    url = cell(row, f"image_url_{n}")
                    if url:
                        # Security fix #6: validate image URL scheme before storing
                        url_str = str(url).strip()
                        parsed = urlparse(url_str)
                        if parsed.scheme in ("http", "https") and parsed.netloc:
                            ProductImage.objects.create(product=product, url=url_str, order=n - 1)

                for n in range(1, 6):
                    lbl = cell(row, f"size_{n}_label")
                    lbl_ka = cell(row, f"size_{n}_label_ka")
                    lbl_ru = cell(row, f"size_{n}_label_ru")
                    sku = cell(row, f"size_{n}_sku")
                    stock = cell(row, f"size_{n}_stock")
                    ready = yn(cell(row, f"size_{n}_ready_to_ship"))
                    pr_usd = cell(row, f"size_{n}_price_usd") or cell(row, f"size_{n}_price")
                    pr_gel = cell(row, f"size_{n}_price_gel")
                    sale_usd = cell(row, f"size_{n}_sale_usd")
                    sale_gel = cell(row, f"size_{n}_sale_gel")
                    if lbl and pr_usd:
                        try:
                            SizeVariant.objects.create(
                                product=product,
                                label=str(lbl),
                                label_ka=str(lbl_ka or ""),
                                label_ru=str(lbl_ru or ""),
                                sku=str(sku).strip() or None,
                                stock=int(stock) if stock not in (None, "") else None,
                                is_ready_to_ship=ready,
                                price_usd=D(str(pr_usd)),
                                price_gel=D(str(pr_gel)) if pr_gel else None,
                                sale_price_usd=D(str(sale_usd)) if sale_usd else None,
                                sale_price_gel=D(str(sale_gel)) if sale_gel else None,
                                sort_order=n - 1,
                            )
                        except Exception:
                            pass
            except Exception as e:
                errors.append({"row": i, "error": str(e)})

        return Response({"created": created_count, "updated": updated_count, "errors": errors}, status=status.HTTP_200_OK)


# ── Catalog filter visibility ─────────────────────────────────────────────────

class AdminCatalogFilterConfigView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.products.models import CatalogFilterConfig, DEFAULT_VISIBLE_FILTERS
        from apps.products.filter_config import _apply_config

        merged_global = dict(DEFAULT_VISIBLE_FILTERS)
        categories: dict[str, dict] = {}
        for cfg in CatalogFilterConfig.objects.filter(vendor__isnull=True):
            if cfg.scope == "global":
                merged_global = _apply_config(merged_global, cfg)
            elif cfg.scope == "category" and cfg.category_slug:
                base = dict(DEFAULT_VISIBLE_FILTERS)
                categories[cfg.category_slug] = _apply_config(base, cfg)
        return Response({"global": merged_global, "categories": categories})

    def patch(self, request):
        from apps.products.models import CatalogFilterConfig

        scope = request.data.get("scope")
        category_slug = (request.data.get("category_slug") or "").strip()
        visible_filters = request.data.get("visible_filters", {})
        if scope not in ("global", "category"):
            return Response({"detail": "Invalid scope."}, status=status.HTTP_400_BAD_REQUEST)
        if scope == "category" and not category_slug:
            return Response({"detail": "category_slug required."}, status=status.HTTP_400_BAD_REQUEST)

        cfg, _ = CatalogFilterConfig.objects.update_or_create(
            scope=scope,
            category_slug=category_slug if scope == "category" else "",
            vendor=None,
            defaults={"visible_filters": visible_filters},
        )
        return Response(cfg.resolved_filters())


# ── Vendor ops (superadmin) ───────────────────────────────────────────────────

class AdminVendorOpsView(APIView):
    permission_classes = [IsAdminUser]

    def patch(self, request, slug):
        from apps.vendors.models import Vendor
        from apps.vendors.serializers import VendorSerializer
        try:
            vendor = Vendor.objects.get(slug=slug)
        except Vendor.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        allowed = (
            "payment_email", "gift_wrap_price_gel", "gift_wrap_price_usd",
            "shipping_email_subject", "shipping_email_body",
        )
        for field in allowed:
            if field in request.data:
                setattr(vendor, field, request.data[field])
        vendor.save()
        log_action(request.user, "vendor_ops_update", "Vendor", vendor.pk, {"vendor_slug": slug})
        return Response(VendorSerializer(vendor).data)


# ── Vendor bulk sale ──────────────────────────────────────────────────────────

class AdminVendorBulkSaleView(APIView):
    """Apply a percentage sale to all active SizeVariants of a vendor's products."""
    permission_classes = [IsAdminUser]

    def post(self, request, slug):
        from apps.vendors.models import Vendor
        from apps.products.models import SizeVariant, Product
        from decimal import Decimal, ROUND_HALF_UP

        try:
            vendor = Vendor.objects.get(slug=slug)
        except Vendor.DoesNotExist:
            return Response({"detail": "Vendor not found."}, status=status.HTTP_404_NOT_FOUND)

        discount_pct = request.data.get("discount_pct")
        currency = request.data.get("currency", "both")  # "GEL", "USD", or "both"

        if discount_pct is None:
            return Response({"detail": "discount_pct is required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            pct = Decimal(str(discount_pct))
            if pct <= 0 or pct >= 100:
                raise ValueError
        except (ValueError, Exception):
            return Response({"detail": "discount_pct must be between 1 and 99."}, status=status.HTTP_400_BAD_REQUEST)

        multiplier = (Decimal("100") - pct) / Decimal("100")

        products = Product.objects.filter(vendor=vendor, status="active")
        updated_count = 0
        for product in products:
            product.is_sale = True
            product.save(update_fields=["is_sale"])
            svs = SizeVariant.objects.filter(product=product, is_active=True)
            for sv in svs:
                if currency in ("USD", "both") and sv.price_usd:
                    sv.sale_price_usd = (sv.price_usd * multiplier).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                if currency in ("GEL", "both") and sv.price_gel:
                    sv.sale_price_gel = (sv.price_gel * multiplier).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                sv.save(update_fields=["sale_price_usd", "sale_price_gel"])
                updated_count += 1

        log_action(request.user, "vendor_bulk_sale", "Vendor", vendor.pk, {
            "vendor": vendor.name, "discount_pct": str(pct), "currency": currency,
            "variants_updated": updated_count,
        })
        return Response({"detail": f"Sale applied to {updated_count} variants.", "variants_updated": updated_count})

    def delete(self, request, slug):
        """Remove sale from all vendor products."""
        from apps.vendors.models import Vendor
        from apps.products.models import SizeVariant, Product

        try:
            vendor = Vendor.objects.get(slug=slug)
        except Vendor.DoesNotExist:
            return Response({"detail": "Vendor not found."}, status=status.HTTP_404_NOT_FOUND)

        products = Product.objects.filter(vendor=vendor)
        for product in products:
            product.is_sale = False
            product.save(update_fields=["is_sale"])
            SizeVariant.objects.filter(product=product).update(sale_price_usd=None, sale_price_gel=None)

        return Response({"detail": "Sale removed from all vendor products."})


# ── Page sections CMS ─────────────────────────────────────────────────────────

class AdminPageSectionListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.cms.models import PageSection
        from apps.cms.serializers import PageSectionSerializer
        from apps.cms.defaults import ensure_page_section_defaults
        ensure_page_section_defaults()
        page = request.query_params.get("page")
        qs = PageSection.objects.all().order_by("page", "sort_order")
        if page:
            qs = qs.filter(page=page)
        return Response(PageSectionSerializer(qs, many=True).data)

    def post(self, request):
        from apps.cms.models import PageSection
        from apps.cms.serializers import PageSectionSerializer
        ser = PageSectionSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        section = ser.save()
        log_action(request.user, "create", "page_section", section.pk, {
            "section_key": section.section_key,
            "page": section.page,
        })
        return Response(PageSectionSerializer(section).data, status=status.HTTP_201_CREATED)


class AdminPageSectionDetailView(APIView):
    permission_classes = [IsAdminUser]

    def patch(self, request, pk):
        from apps.cms.models import PageSection
        from apps.cms.serializers import PageSectionSerializer
        try:
            section = PageSection.objects.get(pk=pk)
        except PageSection.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        ser = PageSectionSerializer(section, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        section = ser.save()
        log_action(request.user, "update", "page_section", section.pk, {
            "section_key": section.section_key,
            "page": section.page,
        })
        return Response(PageSectionSerializer(section).data)

    def delete(self, request, pk):
        from apps.cms.models import PageSection
        try:
            section = PageSection.objects.get(pk=pk)
        except PageSection.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        log_action(request.user, "delete", "page_section", section.pk, {
            "section_key": section.section_key,
            "page": section.page,
        })
        section.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Email Templates ─────────────────────────────────────────────────────────────

class AdminContactMessageListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.contact.models import ContactMessage

        data = [
            {
                "id": message.id,
                "reason": message.reason,
                "first_name": message.first_name,
                "last_name": message.last_name,
                "email": message.email,
                "order_number": message.order_number,
                "message": message.message,
                "attachment": request.build_absolute_uri(message.attachment.url) if message.attachment else "",
                "created_at": message.created_at,
            }
            for message in ContactMessage.objects.all()[:250]
        ]
        return Response(data)


class AdminContactMessageDetailView(APIView):
    permission_classes = [IsAdminUser]

    def delete(self, request, pk):
        from apps.contact.models import ContactMessage

        try:
            message = ContactMessage.objects.get(pk=pk)
        except ContactMessage.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        log_action(request.user, "delete", "contact_message", message.pk, {"email": message.email})
        message.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class AdminEmailTemplateListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.emails.models import EmailTemplate
        from apps.emails.serializers import EmailTemplateListSerializer

        qs = EmailTemplate.objects.select_related("vendor").all()

        return Response(EmailTemplateListSerializer(qs, many=True).data)

    def post(self, request):
        from apps.emails.models import EmailTemplate
        from apps.emails.serializers import EmailTemplateSerializer

        ser = EmailTemplateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        tpl = ser.save()
        log_action(request.user, "create", "email_template", tpl.pk, {
            "name": tpl.name, "event_key": tpl.event_key,
        })
        return Response(EmailTemplateSerializer(tpl).data, status=status.HTTP_201_CREATED)


class AdminEmailTemplateDetailView(APIView):
    permission_classes = [IsAdminUser]

    def _get_template(self, pk, user):
        from apps.emails.models import EmailTemplate
        try:
            tpl = EmailTemplate.objects.select_related("vendor").get(pk=pk)
        except EmailTemplate.DoesNotExist:
            return None
        return tpl

    def get(self, request, pk):
        from apps.emails.serializers import EmailTemplateSerializer
        tpl = self._get_template(pk, request.user)
        if not tpl:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(EmailTemplateSerializer(tpl).data)

    def patch(self, request, pk):
        from apps.emails.serializers import EmailTemplateSerializer
        tpl = self._get_template(pk, request.user)
        if not tpl:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        ser = EmailTemplateSerializer(tpl, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        tpl = ser.save()
        log_action(request.user, "update", "email_template", tpl.pk, {
            "name": tpl.name, "event_key": tpl.event_key,
        })
        return Response(EmailTemplateSerializer(tpl).data)

    def delete(self, request, pk):
        tpl = self._get_template(pk, request.user)
        if not tpl:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        log_action(request.user, "delete", "email_template", tpl.pk, {
            "name": tpl.name, "event_key": tpl.event_key,
        })
        tpl.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class AdminEmailLogListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.emails.models import EmailLog
        from apps.emails.serializers import EmailLogSerializer

        qs = EmailLog.objects.select_related("template").all()[:100]
        return Response(EmailLogSerializer(qs, many=True).data)


class AdminEmailTemplateSeedView(APIView):
    """Install branded default platform email templates."""
    permission_classes = [IsAdminUser]

    def post(self, request):
        from apps.emails.default_templates import install_default_templates

        overwrite = bool(request.data.get("overwrite", False))
        result = install_default_templates(overwrite=overwrite)
        log_action(request.user, "seed", "email_template", "platform", result)
        return Response(result)


class AdminAuctionSubscriberListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.auctions.models import AuctionSubscriber
        qs = AuctionSubscriber.objects.all().order_by("-subscribed_at")
        data = [
            {
                "id": s.id,
                "email": s.email,
                "is_active": s.is_active,
                "subscribed_at": s.subscribed_at,
            }
            for s in qs
        ]
        return Response(data)

# ── Content creators ──────────────────────────────────────────────────────────

class AdminCreatorApplicationListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.creators.models import CreatorApplication
        from apps.creators.serializers import CreatorApplicationSerializer
        status_filter = request.query_params.get("status")
        qs = CreatorApplication.objects.select_related("user", "reviewed_by").all()
        if status_filter:
            qs = qs.filter(status=status_filter)
        return Response(CreatorApplicationSerializer(qs[:200], many=True).data)


class AdminCreatorApplicationDetailView(APIView):
    permission_classes = [IsAdminUser]

    def patch(self, request, pk):
        from django.utils import timezone
        from apps.creators.models import CreatorApplication, ContentCreator
        from apps.creators.serializers import CreatorApplicationSerializer

        try:
            app = CreatorApplication.objects.select_related("user").get(pk=pk)
        except CreatorApplication.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        new_status = request.data.get("status")
        if new_status not in (
            CreatorApplication.STATUS_APPROVED,
            CreatorApplication.STATUS_REJECTED,
            CreatorApplication.STATUS_PENDING,
        ):
            return Response({"detail": "Invalid status."}, status=status.HTTP_400_BAD_REQUEST)

        app.status = new_status
        app.admin_note = request.data.get("admin_note", app.admin_note)
        app.reviewed_by = request.user
        app.reviewed_at = timezone.now()
        app.save()

        if new_status == CreatorApplication.STATUS_APPROVED:
            # Propagate country from application to ContentCreator
            creator_obj, _ = ContentCreator.objects.update_or_create(
                user=app.user,
                defaults={"is_active": True},
            )
            if app.country and not creator_obj.country:
                creator_obj.country = app.country
                creator_obj.save(update_fields=["country"])

        log_action(request.user, "creator_application_review", "CreatorApplication", app.pk, {
            "status": new_status,
            "user": app.user.email,
        })
        return Response(CreatorApplicationSerializer(app).data)


class AdminContentCreatorListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.creators.models import ContentCreator
        from apps.creators.serializers import ContentCreatorSerializer
        qs = ContentCreator.objects.select_related("user", "promo").all()
        return Response(ContentCreatorSerializer(qs, many=True).data)

    def post(self, request):
        """Assign / update creator voucher: user_id, code, discount_percent."""
        from decimal import Decimal
        from apps.users.models import User
        from apps.promo.models import PromoCode
        from apps.creators.models import ContentCreator
        from apps.creators.serializers import ContentCreatorSerializer, AdminAssignVoucherSerializer

        ser = AdminAssignVoucherSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        try:
            user = User.objects.get(pk=data["user_id"])
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

        code = data["code"].strip().upper()
        # Security fix #8: correct voucher collision check with Q objects
        if PromoCode.objects.filter(code=code).exclude(Q(owner=user) | Q(owner__isnull=True)).exists():
            return Response({"detail": "Code already owned by another user."}, status=400)

        promo, _ = PromoCode.objects.update_or_create(
            code=code,
            defaults={
                "owner": user,
                "discount_type": "percent",
                "discount_value": data["discount_percent"],
                "is_active": data.get("is_active", True),
                "max_uses": None,
                "max_uses_per_user": None,
            },
        )
        # Ensure owner set if code existed without owner
        if promo.owner_id != user.id:
            promo.owner = user
            promo.discount_type = "percent"
            promo.discount_value = data["discount_percent"]
            promo.is_active = data.get("is_active", True)
            promo.save()

        creator, _ = ContentCreator.objects.update_or_create(
            user=user,
            defaults={"is_active": True, "promo": promo},
        )
        if creator.promo_id != promo.id:
            creator.promo = promo
            creator.is_active = True
            creator.save()

        log_action(request.user, "creator_voucher_assign", "ContentCreator", creator.pk, {
            "user": user.email,
            "code": code,
            "percent": str(data["discount_percent"]),
        })
        return Response(ContentCreatorSerializer(creator).data, status=status.HTTP_201_CREATED)


class AdminContentCreatorDetailView(APIView):
    """Edit commission % or soft-deactivate a creator. Ledger history is never touched."""
    permission_classes = [IsAdminUser]

    def patch(self, request, pk):
        from apps.creators.models import ContentCreator
        from apps.creators.serializers import ContentCreatorSerializer
        try:
            creator = ContentCreator.objects.select_related("promo", "user").get(pk=pk)
        except ContentCreator.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        new_percent = request.data.get("discount_percent")
        new_active = request.data.get("is_active")
        new_country = request.data.get("country")

        if new_percent is not None and creator.promo:
            from decimal import Decimal as D
            creator.promo.discount_value = D(str(new_percent))
            creator.promo.save(update_fields=["discount_value"])
            log_action(request.user, "creator_percent_update", "ContentCreator", creator.pk, {
                "user": creator.user.email, "new_percent": str(new_percent)
            })

        if new_active is not None:
            creator.is_active = bool(new_active)
            if creator.promo:
                creator.promo.is_active = bool(new_active)
                creator.promo.save(update_fields=["is_active"])
            creator.save(update_fields=["is_active"])

        if new_country is not None:
            # Security fix #19: validate country code against ISO 3166-1 alpha-2
            country_upper = str(new_country).upper().strip()
            if country_upper and country_upper not in VALID_COUNTRY_CODES:
                return Response({"detail": f"Invalid country code: {country_upper}"}, status=status.HTTP_400_BAD_REQUEST)
            creator.country = country_upper
            creator.save(update_fields=["country"])

        creator.refresh_from_db()
        return Response(ContentCreatorSerializer(creator).data)

    def delete(self, request, pk):
        """Soft-deactivate: sets is_active=False on creator+promo+user account. Ledger entries preserved."""
        from apps.creators.models import ContentCreator
        from apps.creators.serializers import ContentCreatorSerializer
        try:
            creator = ContentCreator.objects.select_related("promo", "user").get(pk=pk)
        except ContentCreator.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        # Deactivate creator record
        creator.is_active = False
        creator.save(update_fields=["is_active"])

        # Deactivate associated promo/voucher
        if creator.promo:
            creator.promo.is_active = False
            creator.promo.save(update_fields=["is_active"])

        # Also deactivate the user account so they can no longer log in
        deactivate_user = request.data.get("deactivate_user", True)
        user_deactivated = False
        if deactivate_user and creator.user.is_active:
            creator.user.is_active = False
            creator.user.save(update_fields=["is_active"])
            user_deactivated = True

        log_action(request.user, "creator_deactivated", "ContentCreator", creator.pk, {
            "user": creator.user.email,
            "user_account_deactivated": user_deactivated,
        })
        detail = "Creator deactivated. Ledger history preserved."
        if user_deactivated:
            detail += " User account has also been deactivated."
        return Response({"detail": detail})


class AdminCreatorAcceptedWithoutVoucherView(APIView):
    """Returns approved creator applications whose user does not yet have an active ContentCreator with a promo."""
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.creators.models import CreatorApplication, ContentCreator
        # Users who have a ContentCreator with a promo already assigned
        assigned_user_ids = set(
            ContentCreator.objects.filter(is_active=True, promo__isnull=False)
            .values_list("user_id", flat=True)
        )
        apps = (
            CreatorApplication.objects.filter(status=CreatorApplication.STATUS_APPROVED)
            .select_related("user")
            .exclude(user_id__in=assigned_user_ids)
            .order_by("-created_at")
        )
        return Response([
            {
                "user_id": str(app.user_id),
                "user_email": app.user.email,
                "user_name": app.user.name or app.user.email,
                "country": app.country,
            }
            for app in apps
        ])


class AdminCreatorLedgerListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.creators.models import CreatorLedgerEntry
        from apps.creators.serializers import CreatorLedgerEntrySerializer
        qs = CreatorLedgerEntry.objects.select_related("creator__user", "order").all()
        creator_id = request.query_params.get("creator_id")
        order_number = request.query_params.get("order_number")
        entry_type = request.query_params.get("entry_type")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        if creator_id:
            qs = qs.filter(creator_id=creator_id)
        if order_number:
            qs = qs.filter(order_number__icontains=order_number)
        if entry_type:
            qs = qs.filter(entry_type=entry_type)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        data = []
        for entry in qs[:500]:
            row = CreatorLedgerEntrySerializer(entry).data
            row["creator_email"] = entry.creator.user.email
            row["creator_id"] = entry.creator_id
            data.append(row)
        return Response(data)


class AdminCreatorVoucherUsesView(APIView):
    """All orders that used a creator-owned voucher (even before paid credit)."""
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.creators.models import ContentCreator
        from apps.creators.services import list_voucher_redemptions

        creators = ContentCreator.objects.filter(
            is_active=True, promo__isnull=False
        ).select_related("user", "promo")
        creator_id = request.query_params.get("creator_id")
        credited_filter = request.query_params.get("credited")
        order_status_filter = request.query_params.get("order_status")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        search = request.query_params.get("search", "").strip()

        if creator_id:
            creators = creators.filter(pk=creator_id)

        rows = []
        for creator in creators:
            for item in list_voucher_redemptions(creator.promo, limit=200):
                item["creator_id"] = creator.id
                item["creator_email"] = creator.user.email
                item["voucher_code"] = creator.promo.code if creator.promo_id else None
                item["voucher_percent"] = (
                    str(creator.promo.discount_value) if creator.promo_id else None
                )
                rows.append(item)

        # Apply filters
        if credited_filter is not None:
            want_credited = credited_filter.lower() in ("true", "1", "yes")
            rows = [r for r in rows if r.get("credited") == want_credited]
        if order_status_filter:
            rows = [r for r in rows if r.get("order_status") == order_status_filter]
        if date_from:
            rows = [r for r in rows if (r.get("used_at") or "") >= date_from]
        if date_to:
            rows = [r for r in rows if (r.get("used_at") or "") <= date_to + "T23:59:59"]
        if search:
            sl = search.lower()
            rows = [r for r in rows if sl in (r.get("buyer_email") or "").lower()
                    or sl in (r.get("order_number") or "").lower()]

        rows.sort(key=lambda r: r.get("used_at") or "", reverse=True)
        return Response(rows[:500])


class AdminCreatorPayoutListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.creators.models import CreatorPayoutRequest
        from apps.creators.serializers import CreatorPayoutRequestSerializer
        qs = CreatorPayoutRequest.objects.select_related("creator__user", "processed_by").all()
        status_filter = request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter)
        return Response(CreatorPayoutRequestSerializer(qs[:200], many=True).data)


class AdminCreatorPayoutDetailView(APIView):
    permission_classes = [IsAdminUser]

    def patch(self, request, pk):
        from django.utils import timezone
        from django.db import transaction
        from apps.creators.models import CreatorPayoutRequest, CreatorLedgerEntry
        from apps.creators.serializers import CreatorPayoutRequestSerializer

        try:
            payout = CreatorPayoutRequest.objects.select_related("creator").get(pk=pk)
        except CreatorPayoutRequest.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        new_status = request.data.get("status")
        if new_status not in (CreatorPayoutRequest.STATUS_PAID, CreatorPayoutRequest.STATUS_REJECTED):
            return Response({"detail": "status must be paid or rejected."}, status=400)
        if payout.status != CreatorPayoutRequest.STATUS_PENDING:
            return Response({"detail": "Payout already processed."}, status=400)

        with transaction.atomic():
            payout.status = new_status
            payout.admin_note = request.data.get("admin_note", payout.admin_note)
            payout.processed_by = request.user
            payout.processed_at = timezone.now()
            payout.save()

            hold = CreatorLedgerEntry.objects.filter(
                payout_request=payout,
                entry_type=CreatorLedgerEntry.TYPE_PAYOUT_HOLD,
            ).first()

            if new_status == CreatorPayoutRequest.STATUS_PAID:
                if hold:
                    hold.entry_type = CreatorLedgerEntry.TYPE_PAYOUT_PAID
                    hold.note = f"Payout #{payout.id} marked paid"
                    hold.save(update_fields=["entry_type", "note"])
            else:
                # Reject: remove hold so balance returns
                if hold:
                    hold.delete()

        log_action(request.user, "creator_payout_" + new_status, "CreatorPayoutRequest", payout.pk, {
            "amount": str(payout.amount),
            "creator": payout.creator.user.email,
        })
        return Response(CreatorPayoutRequestSerializer(payout).data)


class AdminCreatorPayoutMinimumView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.creators.services import get_payout_minimum_gel
        return Response({"creator_payout_minimum_gel": str(get_payout_minimum_gel())})

    def patch(self, request):
        from apps.creators.services import set_payout_minimum_gel, get_payout_minimum_gel
        value = request.data.get("creator_payout_minimum_gel")
        if value is None:
            return Response({"detail": "creator_payout_minimum_gel required."}, status=400)
        amount = set_payout_minimum_gel(value)
        log_action(request.user, "creator_payout_minimum_update", "SiteSettings", "global", {
            "creator_payout_minimum_gel": str(amount),
        })
        return Response({"creator_payout_minimum_gel": str(get_payout_minimum_gel())})
